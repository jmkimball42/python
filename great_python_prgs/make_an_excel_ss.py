# Import the `openpyxl` library
import openpyxl
# Create a new Excel workbook
workbook = openpyxl.Workbook()
# Get the active sheet in the workbook (which will be the only sheet at this point)
sheet = workbook.active
# Add some data to the sheet
sheet["A1"] = "Product"
sheet["B1"] = "Sales"
# Add some more data to the sheet
products = ["Product A", "Product B", "Product C"]
sales = [100, 200, 300]
for i in range(len(products)):
    sheet.cell(row=i+2, column=1).value = products[i]
    sheet.cell(row=i+2, column=2).value = sales[i]
# Save the workbook to a file
workbook.save("sales_report.xlsx")